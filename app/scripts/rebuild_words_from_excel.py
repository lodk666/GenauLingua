#!/usr/bin/env python3
"""
Rebuild words table from Excel files (A1/A2/B1).

Usage (inside docker):
  docker compose run --rm app python app/scripts/rebuild_words_from_excel.py /app/Wordsbase
  docker compose run --rm app python app/scripts/rebuild_words_from_excel.py "/app/Wordsbase/А1 v3.xlsx"

Env:
  DATABASE_URL=postgresql+asyncpg://genau_user:genau_pass@postgres:5432/genaulingua_db

What it does (default):
  - TRUNCATE user progress + quiz tables + words (RESTART IDENTITY)
  - Import all Excel files into `words`

If you want "import only, without truncating":
  add flag: --no-truncate
"""

from __future__ import annotations

import argparse
import os
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

from dotenv import load_dotenv
load_dotenv()

from openpyxl import load_workbook

from sqlalchemy import text
from sqlalchemy.ext.asyncio import AsyncSession, create_async_engine
from sqlalchemy.orm import sessionmaker

# Default path: Wordsbase in project root (not inside app/)
_PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
DEFAULT_WORDSBASE = str(_PROJECT_ROOT / "Wordsbase")


# -----------------------------
# Config / aliases
# -----------------------------

COLUMN_ALIASES = {
    # German word
    "word_de": "word_de",
    "wort": "word_de",
    "wort_de": "word_de",
    "de": "word_de",
    "deutsch": "word_de",
    "german": "word_de",
    "нем": "word_de",
    "німецька": "word_de",
    "німецьке": "word_de",
    "слово": "word_de",
    "слово_de": "word_de",
    "слово_deutsch": "word_de",
    "слово_german": "word_de",

    # Article
    "article": "article",
    "artikel": "article",
    "der/die/das": "article",
    "артикль": "article",

    # Part of speech
    "pos": "pos",
    "partofspeech": "pos",
    "part_of_speech": "pos",
    "wortart": "pos",
    "часть_речи": "pos",
    "частина_мови": "pos",

    # Level
    "level": "level",
    "cefr": "level",
    "cefrlevel": "level",
    "niveau": "level",
    "уровень": "level",
    "рівень": "level",

    # Translations
    "translation_ru": "translation_ru",
    "ru": "translation_ru",
    "russian": "translation_ru",
    "перевод_ru": "translation_ru",
    "перевод": "translation_ru",
    "переклад_ru": "translation_ru",

    "translation_uk": "translation_uk",
    "uk": "translation_uk",
    "ua": "translation_uk",
    "ukrainian": "translation_uk",
    "перевод_uk": "translation_uk",
    "переклад": "translation_uk",
    "переклад_uk": "translation_uk",

    # English
    "translation_en": "translation_en",
    "en": "translation_en",
    "english": "translation_en",

    # Turkish
    "translation_tr": "translation_tr",
    "translation_tk": "translation_tr",  # alias if someone used tk
    "tr": "translation_tr",
    "tk": "translation_tr",
    "turkish": "translation_tr",
    "türkisch": "translation_tr",

    # Examples
    "example_de": "example_de",
    "beispiel": "example_de",
    "пример_de": "example_de",
    "приклад_de": "example_de",

    "example_ru": "example_ru",
    "пример_ru": "example_ru",

    "example_uk": "example_uk",
    "приклад_uk": "example_uk",
    "приклад": "example_uk",

    "example_en": "example_en",
    "example_tk": "example_tr",  # alias if someone used tk
    "example_tr": "example_tr",

    # Categories
    "categories": "categories",
    "category": "categories",
    "kategorie": "categories",
    "категории": "categories",
    "категорії": "categories",
}


POS_MAP = {
    "noun": "NOUN", "nomen": "NOUN", "substantiv": "NOUN", "n": "NOUN",
    "сущ": "NOUN", "іменник": "NOUN", "noun.": "NOUN",
    "verb": "VERB", "v": "VERB", "дієслово": "VERB", "глагол": "VERB", "verb.": "VERB",
    "adj": "ADJECTIVE", "adjective": "ADJECTIVE", "adjektiv": "ADJECTIVE", "a": "ADJECTIVE",
    "прилагательное": "ADJECTIVE", "прикметник": "ADJECTIVE", "adj.": "ADJECTIVE",
    "adv": "ADVERB", "adverb": "ADVERB", "adverbium": "ADVERB",
    "наречие": "ADVERB", "прислівник": "ADVERB", "adv.": "ADVERB",
    "pron": "PRONOUN", "pronoun": "PRONOUN", "местоимение": "PRONOUN", "займенник": "PRONOUN",
    "prep": "PREPOSITION", "preposition": "PREPOSITION", "präposition": "PREPOSITION",
    "предлог": "PREPOSITION", "прийменник": "PREPOSITION",
    "conj": "CONJUNCTION", "conjunction": "CONJUNCTION", "союз": "CONJUNCTION", "сполучник": "CONJUNCTION",
    "phrase": "PHRASE", "фраза": "PHRASE", "вираз": "PHRASE",
    "other": "OTHER",
}


LEVEL_RE = re.compile(r"(A1|A2|B1|B2|C1|C2)", re.IGNORECASE)


@dataclass
class WordRow:
    word_de: str
    article: str | None
    pos: str
    level: str
    translation_ru: str | None
    translation_uk: str | None
    translation_en: str | None
    translation_tr: str | None
    example_de: str | None
    example_ru: str | None
    example_uk: str | None
    example_en: str | None
    example_tr: str | None
    categories: list[str]


def _s(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip()


def _normalize_header_cell(v: Any) -> str:
    t = _s(v).lower()
    t = t.replace("\n", " ").replace("\t", " ")
    t = re.sub(r"\s+", " ", t).strip()
    t = t.replace(" ", "_")
    return t


def _detect_level_from_filename(path: Path) -> str | None:
    m = LEVEL_RE.search(path.name.replace("А", "A").replace("В", "B"))
    if not m:
        return None
    return m.group(1).upper()


def _parse_categories(raw: Any) -> list[str]:
    t = _s(raw)
    if not t:
        return []
    sep = ";" if ";" in t else ","
    items = [x.strip() for x in t.split(sep)]
    return [x for x in items if x]


def _map_pos(raw: Any) -> str:
    t = _s(raw).strip()
    if not t:
        return "OTHER"
    key = t.lower().replace(".", "").strip()
    upper = key.upper()
    if upper in {"NOUN", "VERB", "ADJECTIVE", "ADVERB", "PHRASE", "PRONOUN", "PREPOSITION", "CONJUNCTION", "OTHER"}:
        return upper
    if upper == "ADJ":
        return "ADJECTIVE"
    if upper == "ADV":
        return "ADVERB"
    return POS_MAP.get(key, POS_MAP.get(key[:3], "OTHER"))


def _build_header_map(header_row: list[Any]) -> dict[str, int]:
    mapping: dict[str, int] = {}
    for idx, cell_val in enumerate(header_row):
        h = _normalize_header_cell(cell_val)
        if not h:
            continue
        h = re.sub(r"[(){}\[\]:]", "", h).strip("_")
        if h in COLUMN_ALIASES:
            key = COLUMN_ALIASES[h]
            mapping.setdefault(key, idx)
    return mapping


def read_xlsx_rows(xlsx_path: Path, forced_level: str | None = None) -> list[WordRow]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    rows_iter = ws.iter_rows(values_only=True)

    header = None
    for r in rows_iter:
        if any(_s(x) for x in r):
            header = list(r)
            break

    if header is None:
        raise RuntimeError(f"{xlsx_path.name}: empty sheet")

    header_map = _build_header_map(header)

    if "word_de" not in header_map:
        raise RuntimeError(
            f"{xlsx_path.name}: header must contain German word column.\n"
            f"Detected header cells: {[_s(x) for x in header]}"
        )

    file_level = forced_level or _detect_level_from_filename(xlsx_path)

    out: list[WordRow] = []
    for r in rows_iter:
        if r is None:
            continue
        if not any(_s(x) for x in r):
            continue

        word_de = _s(r[header_map["word_de"]])
        if not word_de:
            continue

        level = file_level
        if "level" in header_map:
            lv = _s(r[header_map["level"]]).upper()
            if lv:
                level = lv

        if not level:
            raise RuntimeError(f"{xlsx_path.name}: cannot detect level")

        pos_raw = r[header_map["pos"]] if "pos" in header_map else ""
        pos = _map_pos(pos_raw)

        article = _s(r[header_map["article"]]) if "article" in header_map else ""
        article = article or None

        def get_field(key):
            val = _s(r[header_map[key]]) if key in header_map else ""
            return val or None

        out.append(WordRow(
            word_de=word_de,
            article=article,
            pos=pos,
            level=level,
            translation_ru=get_field("translation_ru"),
            translation_uk=get_field("translation_uk"),
            translation_en=get_field("translation_en"),
            translation_tr=get_field("translation_tr"),
            example_de=get_field("example_de"),
            example_ru=get_field("example_ru"),
            example_uk=get_field("example_uk"),
            example_en=get_field("example_en"),
            example_tr=get_field("example_tr"),
            categories=_parse_categories(r[header_map["categories"]]) if "categories" in header_map else [],
        ))

    return out


def iter_excel_files(path: Path) -> list[Path]:
    if path.is_file():
        return [path]
    if not path.exists():
        raise FileNotFoundError(f"Path not found: {path}")
    files = sorted([p for p in path.iterdir() if p.is_file() and p.suffix.lower() == ".xlsx"])
    if not files:
        raise RuntimeError(f"No .xlsx files found in: {path}")
    return files


async def truncate_tables(session: AsyncSession) -> None:
    await session.execute(text("TRUNCATE TABLE user_words CASCADE"))
    await session.execute(text("TRUNCATE TABLE quiz_questions CASCADE"))
    await session.execute(text("TRUNCATE TABLE quiz_sessions CASCADE"))
    await session.execute(text("TRUNCATE TABLE words RESTART IDENTITY CASCADE"))
    await session.commit()


async def import_words(session: AsyncSession, rows: list[WordRow], chunk_size: int = 500) -> int:
    insert_sql = text("""
        INSERT INTO words
          (word_de, article, pos, level,
           translation_ru, translation_uk, translation_en, translation_tr,
           example_de, example_ru, example_uk, example_en, example_tr,
           categories, times_shown, times_correct, created_at)
        VALUES
          (:word_de, :article,
           CAST(:pos AS partofspeech),
           CAST(:level AS cefrlevel),
           :translation_ru, :translation_uk, :translation_en, :translation_tr,
           :example_de, :example_ru, :example_uk, :example_en, :example_tr,
           :categories, 0, 0, NOW())
        ON CONFLICT (word_de, level)
        DO UPDATE SET
          article = EXCLUDED.article,
          pos = EXCLUDED.pos,
          translation_ru = EXCLUDED.translation_ru,
          translation_uk = EXCLUDED.translation_uk,
          translation_en = EXCLUDED.translation_en,
          translation_tr = EXCLUDED.translation_tr,
          example_de = EXCLUDED.example_de,
          example_ru = EXCLUDED.example_ru,
          example_uk = EXCLUDED.example_uk,
          example_en = EXCLUDED.example_en,
          example_tr = EXCLUDED.example_tr,
          categories = EXCLUDED.categories
    """)

    total = 0
    buf: list[dict[str, Any]] = []

    for w in rows:
        buf.append(dict(
            word_de=w.word_de,
            article=w.article,
            pos=w.pos,
            level=w.level,
            translation_ru=w.translation_ru,
            translation_uk=w.translation_uk,
            translation_en=w.translation_en,
            translation_tr=w.translation_tr,
            example_de=w.example_de,
            example_ru=w.example_ru,
            example_uk=w.example_uk,
            example_en=w.example_en,
            example_tr=w.example_tr,
            categories=w.categories,
        ))

        if len(buf) >= chunk_size:
            await session.execute(insert_sql, buf)
            await session.commit()
            total += len(buf)
            buf.clear()

    if buf:
        await session.execute(insert_sql, buf)
        await session.commit()
        total += len(buf)

    return total


async def async_main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("path", nargs="?", default=DEFAULT_WORDSBASE, help="Excel file or directory with .xlsx files")
    parser.add_argument("--no-truncate", action="store_true", help="Do not truncate tables before import")
    args = parser.parse_args()

    db_url = os.getenv("DATABASE_URL")
    if not db_url:
        raise RuntimeError("DATABASE_URL is not set")

    path = Path(args.path)

    engine = create_async_engine(db_url, echo=False)
    Session = sessionmaker(engine, class_=AsyncSession, expire_on_commit=False)

    excel_files = iter_excel_files(path)

    async with Session() as session:
        if not args.no_truncate:
            print("🗑️ TRUNCATE: user_words, quiz_questions, quiz_sessions, words")
            await truncate_tables(session)
        else:
            print("⚠️ no-truncate: existing data will stay (upsert will update duplicates).")

        grand_total = 0
        for xlsx in excel_files:
            level_hint = _detect_level_from_filename(xlsx)
            print(f"📥 Import: {xlsx.name} (level hint: {level_hint or 'n/a'})")
            rows = read_xlsx_rows(xlsx, forced_level=level_hint)
            inserted = await import_words(session, rows)
            print(f"✅ {xlsx.name}: {inserted} rows")
            grand_total += inserted

        print(f"🏁 Done. Total rows processed: {grand_total}")


def main() -> None:
    import asyncio
    asyncio.run(async_main())


if __name__ == "__main__":
    main()